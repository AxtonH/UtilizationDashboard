"""Excel (.xlsx) export of creative time cards.

Builds a styled workbook from a payload the frontend assembles out of data it
already holds (dashboard state + the bulk daily-hours response), so exporting
costs no Odoo queries. Layout language: every section is an explicitly
bordered block sharing one left edge — gauge cards (doughnut + caption),
two-column hours tables, a bordered calendar grid and project tables — so the
sheets read as designed cards rather than loose cells over gridlines.
"""
from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO
from typing import Any, Dict, List, Mapping, Optional, Tuple

from openpyxl import Workbook
from openpyxl.chart import DoughnutChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import cm_to_EMU, pixels_to_EMU

# Palette: Excel-friendly equivalents of the dashboard's Tailwind tints.
C_TEXT = "0F172A"
C_MUTED = "64748B"
C_FAINT = "94A3B8"
C_EMERALD = "D1FAE5"
C_EMERALD_TXT = "047857"
C_AMBER = "FEF3C7"
C_AMBER_TXT = "B45309"
C_ROSE = "FFE4E6"
C_ROSE_TXT = "BE123C"
C_ORANGE = "FFEDD5"
C_ORANGE_TXT = "C2410C"
C_RED = "FEE2E2"
C_RED_TXT = "B91C1C"
C_SLATE_FILL = "F1F5F9"
C_VIOLET_TXT = "6D28D9"
C_TEAL_TXT = "0F766E"
C_BLUE_TXT = "1D4ED8"
C_INDIGO_TXT = "4338CA"
C_ZEBRA = "F8FAFC"

GAUGE_LOGGED_ARC = "F97316"
GAUGE_LOGGED_TRACK = "FECDD3"
GAUGE_BOOKED_ARC = "8B5CF6"
GAUGE_BOOKED_TRACK = "DDD6FE"
GAUGE_NEUTRAL_TRACK = "E2E8F0"
STATUS_ARC = {"rose": "F43F5E", "amber": "F59E0B", "emerald": "10B981"}

GRID = Side(style="thin", color="E2E8F0")   # inner table lines (slate-200)
EDGE = Side(style="thin", color="94A3B8")   # block outlines (slate-400)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

WEEKDAY_HEADERS = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]

COL0 = 2  # first content column (B) — column A is the sheet margin
# General-based formats: optional-decimal formats like 0.## leave a dangling
# "800." on whole numbers. Percent cells store the percent NUMBER (14.4).
HOURS_FMT = 'General"h"'
NUM_FMT = "General"
PCT_FMT = 'General"%"'

GAUGE_DATA_COL = 30  # hidden helper column (AD) holding gauge chart data


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------

def _num(value: Any) -> float:
    try:
        return round(float(value or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def _fmt(value: Any) -> str:
    num = _num(value)
    if num == int(num):
        return str(int(num))
    return f"{num:.2f}".rstrip("0").rstrip(".")


def _status_key(percent: Any) -> str:
    num = _num(percent)
    if num < 50:
        return "rose"
    if num < 75:
        return "amber"
    return "emerald"


def _status_txt(percent: Any) -> str:
    return {"rose": C_ROSE_TXT, "amber": C_AMBER_TXT, "emerald": C_EMERALD_TXT}[_status_key(percent)]


def _safe_sheet_name(name: str, used: set) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "", str(name or "Creative")).strip() or "Creative"
    base = cleaned[:28]
    candidate = cleaned[:31]
    counter = 2
    while candidate.lower() in used:
        candidate = f"{base}({counter})"
        counter += 1
    used.add(candidate.lower())
    return candidate


def _parse_day_date(value: Any) -> Optional[date]:
    try:
        return date.fromisoformat(str(value)[:10])
    except (TypeError, ValueError):
        return None


def _setup_margins(ws, widths: List[float]) -> None:
    ws.column_dimensions["A"].width = 2.6
    ws.row_dimensions[1].height = 10
    for offset, width in enumerate(widths):
        ws.column_dimensions[get_column_letter(COL0 + offset)].width = width


def _border_with(existing: Border, **sides: Side) -> Border:
    return Border(
        left=sides.get("left", existing.left),
        right=sides.get("right", existing.right),
        top=sides.get("top", existing.top),
        bottom=sides.get("bottom", existing.bottom),
    )


def _grid(ws, r1: int, c1: int, r2: int, c2: int) -> None:
    """Thin grid borders on every cell in the region."""
    for row in range(r1, r2 + 1):
        for col in range(c1, c2 + 1):
            ws.cell(row=row, column=col).border = Border(left=GRID, right=GRID, top=GRID, bottom=GRID)


def _outline(ws, r1: int, c1: int, r2: int, c2: int) -> None:
    """Stronger outline around the region, preserving inner borders."""
    for col in range(c1, c2 + 1):
        top = ws.cell(row=r1, column=col)
        top.border = _border_with(top.border, top=EDGE)
        bottom = ws.cell(row=r2, column=col)
        bottom.border = _border_with(bottom.border, bottom=EDGE)
    for row in range(r1, r2 + 1):
        left = ws.cell(row=row, column=c1)
        left.border = _border_with(left.border, left=EDGE)
        right = ws.cell(row=row, column=c2)
        right.border = _border_with(right.border, right=EDGE)


def _banner(ws, row: int, c1: int, c2: int, title: str, hint: str = "") -> None:
    """Section title row: filled band spanning the block width."""
    for col in range(c1, c2 + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=C_SLATE_FILL)
    cell = ws.cell(row=row, column=c1, value=title.upper())
    cell.font = Font(size=9, bold=True, color=C_MUTED)
    cell.alignment = LEFT
    if hint:
        hc = ws.cell(row=row, column=c2, value=hint)
        hc.font = Font(size=8, color=C_FAINT)
        hc.alignment = RIGHT


# ---------------------------------------------------------------------------
# Gauge card: doughnut chart + caption inside one outlined block
# ---------------------------------------------------------------------------

GAUGE_CARD_ROWS = 13   # card height in rows
GAUGE_CHART_CM = 3.4   # doughnut size
_DEFAULT_ROW_PX = 20   # Excel default row height (15pt)


def _col_px(ws, col: int) -> float:
    """Approximate rendered pixel width of a column (Calibri 11)."""
    width = ws.column_dimensions[get_column_letter(col)].width or 8.43
    return width * 7 + 5


def _centered_chart_anchor(ws, r1: int, c1: int, c2: int, zone_rows: int) -> OneCellAnchor:
    """Anchor centering the gauge within the card's chart zone."""
    chart_px = GAUGE_CHART_CM * 96 / 2.54
    card_px = sum(_col_px(ws, col) for col in range(c1, c2 + 1))
    x_off = max((card_px - chart_px) / 2, 0)
    y_off = max((zone_rows * _DEFAULT_ROW_PX - chart_px) / 2, 0)
    marker = AnchorMarker(
        col=c1 - 1, colOff=pixels_to_EMU(int(x_off)),
        row=r1 - 1, rowOff=pixels_to_EMU(int(y_off)),
    )
    size = XDRPositiveSize2D(cm_to_EMU(GAUGE_CHART_CM), cm_to_EMU(GAUGE_CHART_CM))
    return OneCellAnchor(_from=marker, ext=size)


def _gauge_card(ws, r1: int, c1: int, c2: int, label: str, percent: Any,
                arc: str, track: str, caption_color: str, data_row: int) -> int:
    """Card spanning GAUGE_CARD_ROWS rows: chart on top, label + % beneath."""
    r2 = r1 + GAUGE_CARD_ROWS - 1
    value = max(0.0, min(_num(percent), 100.0))
    ws.cell(row=data_row, column=GAUGE_DATA_COL, value=round(value, 1))
    ws.cell(row=data_row + 1, column=GAUGE_DATA_COL, value=round(100 - value, 1))

    chart = DoughnutChart(holeSize=62)
    chart.add_data(
        Reference(ws, min_col=GAUGE_DATA_COL, min_row=data_row, max_row=data_row + 1),
        titles_from_data=False,
    )
    chart.series[0].data_points = [
        DataPoint(idx=0, spPr=GraphicalProperties(solidFill=arc)),
        DataPoint(idx=1, spPr=GraphicalProperties(solidFill=track)),
    ]
    chart.legend = None
    chart.visible_cells_only = False
    chart.width = GAUGE_CHART_CM
    chart.height = GAUGE_CHART_CM
    # Chart zone = card rows above the two caption rows + bottom padding row.
    ws.add_chart(chart, _centered_chart_anchor(ws, r1, c1, c2, GAUGE_CARD_ROWS - 3))

    lab = ws.cell(row=r2 - 2, column=c1, value=label)
    lab.font = Font(size=8, bold=True, color=C_MUTED)
    lab.alignment = CENTER
    ws.merge_cells(start_row=r2 - 2, start_column=c1, end_row=r2 - 2, end_column=c2)
    val = ws.cell(row=r2 - 1, column=c1, value=round(_num(percent), 1))
    val.number_format = PCT_FMT
    val.font = Font(size=14, bold=True, color=caption_color)
    val.alignment = CENTER
    ws.merge_cells(start_row=r2 - 1, start_column=c1, end_row=r2 - 1, end_column=c2)

    _outline(ws, r1, c1, r2, c2)
    return r2


# ---------------------------------------------------------------------------
# Two-column label/value table (hours, totals)
# ---------------------------------------------------------------------------

HOURS_CATEGORIES = [
    ("Available Hours", "available_hours", C_MUTED, C_TEXT),
    ("Base Hours", "base_hours", C_TEAL_TXT, C_TEAL_TXT),
    ("Time Off", "time_off_hours", C_ORANGE_TXT, C_ORANGE_TXT),
    ("Public Holiday", "public_holiday_hours", C_RED_TXT, C_RED_TXT),
    ("Booked Hours", "planned_hours", C_BLUE_TXT, C_BLUE_TXT),
    ("Logged Hours", "logged_hours", C_INDIGO_TXT, C_INDIGO_TXT),
    ("Overtime", "overtime_hours", C_VIOLET_TXT, C_VIOLET_TXT),
]


def _hours_table(ws, r1: int, c1: int, c2: int, title: str, values: Mapping[str, float],
                 pad_to_row: Optional[int] = None) -> int:
    """Banner + seven bordered label/value rows. Returns next free row.

    ``pad_to_row`` extends the block's grid/outline so it can bottom-align
    with neighboring blocks (e.g. the gauge cards on the Overview band).
    """
    split = c1 + (c2 - c1) // 2  # label spans c1..split, value spans split+1..c2
    _banner(ws, r1, c1, c2, title)
    row = r1 + 1
    for label, key, label_color, value_color in HOURS_CATEGORIES:
        lc = ws.cell(row=row, column=c1, value=label)
        lc.font = Font(size=9, bold=True, color=label_color)
        lc.alignment = LEFT
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=split)
        vc = ws.cell(row=row, column=split + 1, value=_num(values.get(key)))
        vc.number_format = HOURS_FMT
        vc.font = Font(size=9, bold=True, color=value_color)
        vc.alignment = RIGHT
        ws.merge_cells(start_row=row, start_column=split + 1, end_row=row, end_column=c2)
        row += 1
    last_row = row - 1
    _grid(ws, r1, c1, last_row, c2)  # grid only the content rows
    if pad_to_row and pad_to_row > last_row:
        last_row = pad_to_row       # padding rows get the outline only
    _outline(ws, r1, c1, last_row, c2)
    return last_row + 2


# ---------------------------------------------------------------------------
# Calendar
# ---------------------------------------------------------------------------

def _day_cell_style(day: Mapping[str, Any], today: date) -> Tuple[Any, str, str]:
    day_date = _parse_day_date(day.get("date"))
    day_num = str(day_date.day) if day_date else ""
    logged = _num(day.get("logged"))
    booked = _num(day.get("booked"))
    overtime = _num(day.get("overtime"))
    expected = _num(day.get("expected"))
    time_off = _num(day.get("time_off"))
    holiday = _num(day.get("holiday"))
    has_activity = logged > 0 or booked > 0 or overtime > 0
    is_future = bool(day_date and day_date > today)

    ot_suffix = f"\n{_fmt(overtime)} OT" if overtime > 0 else ""

    if holiday > 0 and holiday >= expected and not has_activity:
        return f"{day_num}\nPH{ot_suffix}", C_RED, C_RED_TXT
    if time_off > 0 and expected > 0 and time_off >= expected and not has_activity:
        return f"{day_num}\nTO{ot_suffix}", C_ORANGE, C_ORANGE_TXT
    if expected == 0 and not has_activity:
        return (int(day_num) if day_num else ""), C_SLATE_FILL, C_FAINT
    if is_future:
        if booked > 0:
            return f"{day_num}\n-/{_fmt(booked)}", "FFFFFF", C_FAINT
        return (int(day_num) if day_num else ""), "FFFFFF", C_FAINT

    value = f"\n{_fmt(logged)}/{_fmt(booked)}"
    if logged > 0 and logged >= booked - 0.01:
        return f"{day_num}{value}{ot_suffix}", C_EMERALD, C_EMERALD_TXT
    if logged > 0:
        return f"{day_num}{value}{ot_suffix}", C_AMBER, C_AMBER_TXT
    if booked > 0 or expected > 0:
        return f"{day_num}{value}{ot_suffix}", C_ROSE, C_ROSE_TXT
    return f"{day_num}{value}{ot_suffix}", "FFFFFF", C_TEXT


def _write_calendar(ws, start_row: int, days: List[Mapping[str, Any]], today: date) -> int:
    c1, c2 = COL0, COL0 + 6
    _banner(ws, start_row, c1, c2, "Daily Hours")
    row = start_row + 1
    grid_start = start_row

    months: Dict[str, List[Mapping[str, Any]]] = {}
    for day in days:
        months.setdefault(str(day.get("date"))[:7], []).append(day)

    for month_key, month_days in months.items():
        if len(months) > 1:
            first = _parse_day_date(month_days[0].get("date"))
            tc = ws.cell(row=row, column=c1, value=first.strftime("%B %Y") if first else month_key)
            tc.font = Font(size=9, bold=True, color=C_MUTED)
            row += 1

        for offset, label in enumerate(WEEKDAY_HEADERS):
            cell = ws.cell(row=row, column=c1 + offset, value=label)
            cell.font = Font(size=8, bold=True, color=C_FAINT)
            cell.alignment = CENTER
            cell.fill = PatternFill("solid", fgColor=C_ZEBRA)
        row += 1

        first_date = _parse_day_date(month_days[0].get("date"))
        col = c1 + ((first_date.weekday() + 1) % 7 if first_date else 0)  # Sunday-first
        # Leading blanks stay part of the bordered grid.
        for day in month_days:
            value, fill, font_color = _day_cell_style(day, today)
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = CENTER
            cell.font = Font(size=8, color=font_color)
            cell.fill = PatternFill("solid", fgColor=fill)
            ws.row_dimensions[row].height = 30
            col += 1
            if col > c2:
                col = c1
                row += 1
        if col != c1:
            row += 1
        row += 0  # months separated by their own header rows

    _grid(ws, grid_start, c1, row - 1, c2)
    _outline(ws, grid_start, c1, row - 1, c2)

    legend = ws.cell(
        row=row,
        column=c1,
        value="Cells: day / logged/booked · green = logged ≥ booked · amber = under-logged · "
        "red = no hours · TO = time off · PH = PrezHoliday · OT = overtime",
    )
    legend.font = Font(size=8, color=C_FAINT)
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    return row + 2


# ---------------------------------------------------------------------------
# Project tables
# ---------------------------------------------------------------------------

def _project_table(
    ws, start_row: int, title: str, hint: str, rows: List[Mapping[str, Any]],
    col_a: str, col_b: str, key_a: str, key_b: str, value_color: str,
) -> int:
    c1, c2 = COL0, COL0 + 6
    _banner(ws, start_row, c1, c2, title, hint)
    row = start_row + 1

    if not rows:
        cell = ws.cell(row=row, column=c1, value="None this period")
        cell.font = Font(size=9, italic=True, color=C_FAINT)
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        _grid(ws, start_row, c1, row, c2)
        _outline(ws, start_row, c1, row, c2)
        return row + 2

    for col, label, align in ((c1, "Project", LEFT), (c2 - 1, col_a, RIGHT), (c2, col_b, RIGHT)):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = Font(size=8, bold=True, color=C_MUTED)
        cell.alignment = align
        cell.fill = PatternFill("solid", fgColor=C_ZEBRA)
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2 - 2)
    for col in range(c1, c2 + 1):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=C_ZEBRA)
    row += 1

    for entry in rows:
        name_cell = ws.cell(row=row, column=c1, value=str(entry.get("project_name") or "Unassigned Project"))
        name_cell.font = Font(size=9, color=C_TEXT)
        name_cell.alignment = LEFT
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2 - 2)
        for col, key in ((c2 - 1, key_a), (c2, key_b)):
            cell = ws.cell(row=row, column=col, value=_num(entry.get(key)))
            cell.number_format = NUM_FMT
            cell.font = Font(size=9, bold=True, color=value_color)
            cell.alignment = RIGHT
        row += 1

    _grid(ws, start_row, c1, row - 1, c2)
    _outline(ws, start_row, c1, row - 1, c2)
    return row + 1


# ---------------------------------------------------------------------------
# Sheets
# ---------------------------------------------------------------------------

def _sheet_header(ws, creative_or_title: str, subtitle: str, period_label: str, c2: int) -> None:
    name_cell = ws.cell(row=2, column=COL0, value=creative_or_title)
    name_cell.font = Font(size=15, bold=True, color=C_TEXT)
    ws.merge_cells(start_row=2, start_column=COL0, end_row=2, end_column=c2)
    if subtitle:
        sub = ws.cell(row=3, column=COL0, value=subtitle)
        sub.font = Font(size=9, color=C_MUTED)
        ws.merge_cells(start_row=3, start_column=COL0, end_row=3, end_column=c2)
    period = ws.cell(row=4, column=COL0, value=period_label)
    period.font = Font(size=9, color=C_FAINT)
    ws.merge_cells(start_row=4, start_column=COL0, end_row=4, end_column=c2)


def _write_creative_sheet(ws, creative: Mapping[str, Any], period_label: str, today: date) -> None:
    _setup_margins(ws, [13] * 7)
    last_col = COL0 + 6

    org_bits = [
        str(v) for v in (
            creative.get("department"),
            creative.get("business_unit") or creative.get("market_display"),
            creative.get("sub_business_unit"),
            creative.get("pod") or (creative.get("pool_display") if creative.get("pool_display") != "No Pool" else None),
        ) if v
    ]
    _sheet_header(ws, str(creative.get("name") or "Unnamed Creative"), "  ·  ".join(org_bits), period_label, last_col)

    # Two gauge cards side by side (cols B..D and F..H, gutter at E).
    logged_pct = creative.get("logged_utilization")
    booked_pct = creative.get("planned_utilization")
    _gauge_card(ws, 6, COL0, COL0 + 2, "LOGGED UTILIZATION", logged_pct,
                STATUS_ARC[_status_key(logged_pct)], GAUGE_NEUTRAL_TRACK, _status_txt(logged_pct), 2)
    _gauge_card(ws, 6, COL0 + 4, COL0 + 6, "BOOKED UTILIZATION", booked_pct,
                STATUS_ARC[_status_key(booked_pct)], GAUGE_NEUTRAL_TRACK, _status_txt(booked_pct), 5)

    hours_values = {key: _num(creative.get(key)) for _, key, _, _ in HOURS_CATEGORIES}
    row = _hours_table(ws, 6 + GAUGE_CARD_ROWS + 1, COL0, last_col, "Hours", hours_values)

    row = _write_calendar(ws, row, list(creative.get("days") or []), today)

    row = _project_table(
        ws, row, "Worked Projects", "logged / booked",
        list(creative.get("projects") or []),
        "Logged", "Booked", "logged", "booked", C_TEXT,
    )
    _project_table(
        ws, row, "Overtime", "overtime / logged over booked",
        list(creative.get("overtime_projects") or []),
        "Overtime", "Logged", "overtime", "logged_overtime", C_VIOLET_TXT,
    )

    ws.column_dimensions[get_column_letter(GAUGE_DATA_COL)].hidden = True


def _write_overview_sheet(ws, creatives: List[Mapping[str, Any]], sheet_names: List[str], period_label: str) -> None:
    ws.title = "Overview"
    widths = [24, 14, 14, 22, 10, 10, 10, 10, 10, 10]
    _setup_margins(ws, widths)
    last_col = COL0 + len(widths) - 1

    _sheet_header(
        ws, "Creative Time Cards",
        f"{period_label}  ·  {len(creatives)} creatives  ·  exported {datetime.now():%d/%m/%Y %H:%M}",
        "", last_col,
    )

    totals = {key: sum(_num(c.get(key)) for c in creatives) for _, key, _, _ in HOURS_CATEGORIES}
    available = totals["available_hours"]
    team_logged_pct = (totals["logged_hours"] / available * 100.0) if available > 0 else 0.0
    team_booked_pct = (totals["planned_hours"] / available * 100.0) if available > 0 else 0.0

    # One aligned band spanning the roster's width: two gauge cards + the
    # totals table, all bottom-aligned at row 14.
    _gauge_card(ws, 5, COL0, COL0 + 1, "LOGGED UTILIZATION", team_logged_pct,
                GAUGE_LOGGED_ARC, GAUGE_LOGGED_TRACK, GAUGE_LOGGED_ARC, 2)
    _gauge_card(ws, 5, COL0 + 2, COL0 + 3, "BOOKED UTILIZATION", team_booked_pct,
                GAUGE_BOOKED_ARC, GAUGE_BOOKED_TRACK, GAUGE_BOOKED_ARC, 5)
    _hours_table(ws, 5, COL0 + 4, last_col, "Total Hours", totals, pad_to_row=5 + GAUGE_CARD_ROWS - 1)

    # Roster: full bordered grid with zebra rows and a bold TOTAL row.
    headers = ["Creative", "BU", "Sub BU", "Pod", "Available", "Booked", "Logged", "OT", "Booked %", "Logged %"]
    header_row = 5 + GAUGE_CARD_ROWS + 2
    for offset, header_label in enumerate(headers):
        cell = ws.cell(row=header_row, column=COL0 + offset, value=header_label)
        cell.font = Font(size=8, bold=True, color=C_MUTED)
        cell.fill = PatternFill("solid", fgColor=C_SLATE_FILL)
        cell.alignment = LEFT if offset <= 3 else RIGHT

    for idx, (creative, sheet_name) in enumerate(zip(creatives, sheet_names)):
        row = header_row + 1 + idx
        name_cell = ws.cell(row=row, column=COL0, value=str(creative.get("name") or ""))
        name_cell.hyperlink = f"#'{sheet_name}'!A1"
        name_cell.font = Font(size=9, color=C_BLUE_TXT, underline="single")
        for offset, value in enumerate((
            creative.get("business_unit") or creative.get("market_display") or "",
            creative.get("sub_business_unit") or "",
            creative.get("pod") or "",
        ), start=1):
            cell = ws.cell(row=row, column=COL0 + offset, value=value)
            cell.font = Font(size=9, color=C_TEXT)
            cell.alignment = LEFT
        for offset, key in enumerate(("available_hours", "planned_hours", "logged_hours", "overtime_hours"), start=4):
            cell = ws.cell(row=row, column=COL0 + offset, value=_num(creative.get(key)))
            cell.number_format = NUM_FMT
            cell.font = Font(size=9, color=C_TEXT)
            cell.alignment = RIGHT
        for offset, value in enumerate((creative.get("planned_utilization"), creative.get("logged_utilization")), start=8):
            cell = ws.cell(row=row, column=COL0 + offset, value=round(_num(value), 1))
            cell.number_format = PCT_FMT
            cell.font = Font(size=9, color=C_TEXT)
            cell.alignment = RIGHT
        if idx % 2 == 1:
            for offset in range(len(headers)):
                ws.cell(row=row, column=COL0 + offset).fill = PatternFill("solid", fgColor=C_ZEBRA)

    total_row = header_row + 1 + len(creatives)
    label_cell = ws.cell(row=total_row, column=COL0, value="TOTAL")
    label_cell.font = Font(size=9, bold=True, color=C_TEXT)
    for offset, value in enumerate((
        totals["available_hours"], totals["planned_hours"], totals["logged_hours"], totals["overtime_hours"],
    ), start=4):
        cell = ws.cell(row=total_row, column=COL0 + offset, value=value)
        cell.number_format = NUM_FMT
        cell.font = Font(size=9, bold=True, color=C_TEXT)
        cell.alignment = RIGHT
    for offset, value in enumerate((team_booked_pct, team_logged_pct), start=8):
        cell = ws.cell(row=total_row, column=COL0 + offset, value=round(_num(value), 1))
        cell.number_format = PCT_FMT
        cell.font = Font(size=9, bold=True, color=C_TEXT)
        cell.alignment = RIGHT
    for offset in range(len(headers)):
        ws.cell(row=total_row, column=COL0 + offset).fill = PatternFill("solid", fgColor=C_SLATE_FILL)

    _grid(ws, header_row, COL0, total_row, COL0 + len(headers) - 1)
    _outline(ws, header_row, COL0, total_row, COL0 + len(headers) - 1)

    ws.column_dimensions[get_column_letter(GAUGE_DATA_COL)].hidden = True


def build_timecards_workbook(payload: Mapping[str, Any]) -> BytesIO:
    """Build the export workbook and return it as an in-memory stream."""
    creatives = [c for c in (payload.get("creatives") or []) if isinstance(c, Mapping)]
    period_label = str(payload.get("period_label") or payload.get("selected_month") or "")
    today = date.today()

    workbook = Workbook()
    used_names: set = set()
    used_names.add("overview")
    sheet_names: List[str] = []

    for creative in creatives:
        sheet_name = _safe_sheet_name(str(creative.get("name") or "Creative"), used_names)
        sheet_names.append(sheet_name)
        ws = workbook.create_sheet(title=sheet_name)
        _write_creative_sheet(ws, creative, period_label, today)

    _write_overview_sheet(workbook.worksheets[0], creatives, sheet_names, period_label)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream
